import re
import json
import numpy as np
import pandas as pd
import os
from win32com import client
from openpyxl import load_workbook
from openpyxl.styles import Font


from inc.get_coord import get_location, get_info_object, get_locality_population
from inc.population_in_district import count_all_objs_in_region, get_objs_in_district_from_cache, delete_if_older_30_days, get_residence
from inc.population_from_h3 import get_all_objs_from_kontur_population, get_residents_from_cache_h3
from inc.parse_data import parse_from_address_settlement
from inc.get_predicted import get_predicted

def save_opening_output_file(file_path):
    if os.path.exists(file_path):
        excelApp = client.Dispatch('Excel.Application')
        book = excelApp.Workbooks.open(os.path.abspath(file_path))
        excelApp.DisplayAlerts = False
        book.Save()
        book.Close()
#
def parse_info(row):

    lat, lon = get_coord_from_hyper(row[18].value)
    return {
        'region': row[1].value,
        'id': re.sub(r'^.+"([^"]+)"\)$', r'\1', row[3].value) if row[3].value else None,
        'address':row[6].value,
        'price':row[7].value, #float(format(price_m2 / int(population), ".2f"))
        'price_m2':row[8].value if row[8].value else float(format(int(row[7].value)/int(row[2].value), '.2f')),
        'people/m2':row[10].value,
        'entity/m2':row[11].value,
        'residents':row[12].value,
        'people_in_city':row[13].value,
        'entity':row[14].value,
        'postal_distance':row[15].value,
        'cad_num':row[20].value,
        'lat':lat,
        'lon':lon,
        'floor': row[21].value,
    }

def try_get_coords_again(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    df_population_in_locality = pd.read_excel('konturs/Population.xlsx',
                                              usecols=['municipality', 'settlement', 'type', 'population'])  # 'region',

    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        row_info = parse_info(row)

        if row_info['lat'] and row_info['lon']:
            if not os.path.exists(f'cache/tmp_loc/{row_info["id"]}.json'):
                save_address_with_geo(row_info["id"], row_info['lat'], row_info['lon'], row_info['address'], row_info['cad_num'])

        elif not row_info['lat'] or not row_info['lon']:
            print(f"get info {index} in xls")
            info_object = get_info_object(row_info["id"], row_info['address'], row_info['cad_num'])

            if info_object:
                settlement_population = ""
                settlement_type = info_object.get('settlement_type')
                settlement = info_object.get('settlement')
                district = info_object.get('area')
                if settlement_type and settlement:
                    settlement_population = get_locality_population(df=df_population_in_locality, settlement=settlement,
                                                                    settlement_type=settlement_type, area=district)

                lat = info_object['lat']
                lon = info_object['lon']
                coord = f'''=HYPERLINK("https://yandex.ru/maps/?&text={lat}, {lon}", "{lat}, {lon}")''' if lat and lon else ""

                ws[f'S{index+1}'] = coord
                ws[f'S{index+1}'].font = Font(color='0000FF', underline='single')
                ws[f'G{index+1}'] = f'{info_object["address"]}'
                ws[f'N{index+1}'] = settlement_population

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)

def save_population_in_cache(id, postal_distance):
    with open(f'cache/tmp_loc/{id}.json', encoding='utf8') as f:
        info_object = json.load(f)
        info_object[0]['data']['postal_distance'] = postal_distance

        with open(f'cache/tmp_loc/{id}.json', "w", encoding='utf8') as file:
            json.dump(info_object, file, ensure_ascii=False, indent=4)


def save_address_with_geo(id, lat, lon, address, cadnum = None):
    data_json = [{
        "value": f"{address}",
        "unrestricted_value": f"{address}",
        "data": {
            "country": "Россия",
            "country_iso_code": "RU",
            "geo_lat": f"{lat}",
            "geo_lon": f"{lon}",
            "qc_geo": "1",
            "flat_cadnum": f"{cadnum}" if cadnum else None
        }
    }]
    if not os.path.exists('cache/tmp_loc'):
        os.makedirs('cache/tmp_loc')
    with open(f'cache/tmp_loc/{id}.json', "w", encoding='utf8') as file:
        json.dump(data_json, file, ensure_ascii=False, indent=4)


def get_from_kontur_population(file_xlsx_path):
    print('get_from_kontur_population')
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            row_info = parse_info(row)
            lat, lon = row_info["lat"], row_info["lon"]
            if (lat and lon) and (not row_info["residents"]):
                if not os.path.exists(f'cache/population_in_h3/{lat}_{lon}.json'):
                    objs[index] = {"lat": lat, "lon": lon, "price_m2": row_info["price_m2"]}
                else:
                    with open(f'cache/population_in_h3/{lat}_{lon}.json', encoding='utf8') as f:
                        info_object = json.load(f)
                        set_in_cell_h3people(ws, index=index + 1, population=info_object["population"],
                                    price_m2=info_object["price_m2"])

        except Exception as _ex:
            print(_ex, row_info, index)
            raise

    get_all_objs_from_kontur_population(objs)
    for index in objs:
        set_in_cell_h3people(ws, index=index+1, population=objs[index]["population"], price_m2=objs[index]["price_m2"])

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)

def get_coord_from_hyper(hyper):
    coord = re.sub(r'^.+"([^"]+)"\)$', r'\1', hyper).split(',') if hyper else None
    if coord:
        return float(coord[0]), float(coord[1])
    return None, None

def set_in_cell_h3people(ws, index, population, price_m2):
    ws[f'M{index}'] = int(population)
    if price_m2:
        ws[f'K{index}'] = float(format(price_m2 / int(population), ".2f")) if int(population) else ""
        ws[f'K{index}'].number_format = '# ### ##0.0₽'


def set_population_in_xls_from_cache(file_xlsx_path):
    print('set_population_in_xls_from_cache')
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            row_info = parse_info(row)
            if row_info["entity"]: #уже есть данные
                continue
            lat, lon = row_info["lat"], row_info["lon"]
            if not lat or not lon:
                continue

            if os.path.exists(f'cache/objs_in_district/{lat}_{lon}.json'):
                residence, entity = get_residence(lat, lon)
                #entity = get_objs_in_district_from_cache(lat, lon)
                if entity != "":
                    if row_info['residents'] < int(residence):
                        ws[f'M{index+1}'] = int(residence)
                        if row_info["price_m2"]:
                            ws[f'K{index + 1}'] = float(
                                format(row_info["price_m2"] / int(residence), ".2f")) if entity else ""
                            ws[f'K{index + 1}'].number_format = '# ### ##0.0₽'


                    ws[f'O{index+1}'] = int(entity) if entity else ""
                    ws[f'T{index+1}'].font = Font(underline='single', color='0000FF')
                    ws[f'T{index+1}'] = f"""=HYPERLINK("{os.path.abspath("cache/objs_in_district")}/{lat}_{lon}.json", "{lat}_{lon}.json")"""
                    if row_info["price_m2"]:
                        ws[f'L{index+1}'] = float(format(row_info["price_m2"] / int(entity), ".2f")) if entity else ""
                        ws[f'L{index+1}'].number_format = '# ### ##0₽'

                else:
                    delete_if_older_30_days(f'cache/objs_in_district/{lat}_{lon}.json')

        except Exception as _ex:
            print(_ex, row_info, index)
            raise

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)


def get_population_from_osm(file_xlsx_path):
    print('get_population_from_osm')
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            row_info = parse_info(row)
            if row_info['people_in_city']:
                continue

            if not os.path.exists(f'cache/objs_in_district/{row_info["lat"]}_{row_info["lon"]}.json') and (row_info.get("lat") and row_info.get("lon")):
                objs[index] = {"index": index, "lat": row_info["lat"], "lon": row_info["lon"], "price_m2": row_info["price_m2"], 'nodes':set(), 'entity':{}}
                list_obj = objs_by_regions.get(row_info["region"]) or []
                list_obj.insert(index,objs[index])
                objs_by_regions[row_info["region"]] = list_obj

        except Exception as _ex:
            print(_ex, row_info, index)
            raise

    for region in objs_by_regions:
        print(objs_by_regions[region])
        count_all_objs_in_region(objs_by_regions[region], region)

    set_population_in_xls_from_cache(file_xlsx_path)

    return

def set_population_locality(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:

            row_info = parse_info(row)
            if not row_info['people_in_city']:
                print(parse_from_address_settlement(row_info['address']))

        except Exception as _ex:
            print(_ex)
            raise
    return


def main(out_file = 'torgi/output.xlsx'):
    try_get_coords_again(out_file)
    get_from_kontur_population(out_file)
    get_population_from_osm(out_file)
    get_predicted(out_file)
if __name__ == "__main__":
    main()


