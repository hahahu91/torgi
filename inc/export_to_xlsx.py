import json
from win32com import client
import pandas as pd
from pandas import json_normalize
from os import path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def set_predicted_in_xls(file_xlsx_path, df):
    print('set_predicted_in_xls')
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for i, row in df.iterrows():
        ws[f'W{int(row[0]) + 2}'].number_format = '# ##0₽;-# ##0₽'
        ws[f'W{int(row[0]) + 2}'] = row["predicted"]

        ws[f'X{int(row[0]) + 2}'].number_format = '# ##0₽;-# ##0₽'
        ws[f'X{int(row[0]) + 2}'] = row["predicted"] - row["Цена за кв.м"]
        if row["predicted"] - row["Цена за кв.м"] > 100:
            ws[f'P{int(row[0]) + 2}'] = ((row["predicted"] - row["Цена за кв.м"]) * row['Общая площадь'] / row["Цена"])
            ws[f'P{int(row[0]) + 2}'].number_format = '0%'

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)
    return

def install_setting_of_columns(writer):
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    #worksheet.column_dimensions['A'].hidden = True
    worksheet.autofilter('B1:AA1000')
    currency_format = workbook.add_format({'num_format': '# ### ##0₽'})
    population_format = workbook.add_format({'num_format': '# ### ##0'})
    date_forman = workbook.add_format({'num_format': 'dd.mm.yy hh:mm'})
    currency_format_with_penny = workbook.add_format({'num_format': '# ### ##0.0₽'})
    format_area = workbook.add_format({'num_format': '# ##0.0м2'})
    format_distance = workbook.add_format({'num_format': '#\ ##0\ \м'})
    format_hyper = workbook.add_format({'font_color': 'blue', 'underline': True})

    border_format = workbook.add_format({'border': 1})

    center = workbook.add_format().set_center_across()
    hidden = workbook.add_format().set_hidden()
    worksheet.set_column('A:A', 3, hidden)
    # Регион
    worksheet.set_column('B:B', 3, center)
    # area
    worksheet.set_column('C:C', 9, format_area)

    worksheet.set_column('D:D', 23, format_hyper)
    worksheet.set_column('S:T', 13, format_hyper)

    # name
    worksheet.set_column('E:E', 12)

    worksheet.set_column('G:G', 40)
    worksheet.set_column('H:H', 12, currency_format)
    worksheet.set_column('F:F', 12, date_forman)
    worksheet.set_column('M:M', 5)
    worksheet.set_column('N:N', 10, population_format)

    worksheet.set_column('O:O', 5)
    worksheet.set_column('U:U', 12)
    worksheet.set_column('U:U', 5)
    worksheet.set_column('V:V', 3, center)
    worksheet.set_column('P:P', 6)
    worksheet.set_column('Q:R', 3, center)
    worksheet.set_column('Y:AA', 3, center)
    #worksheet.set_column('T:T', 18)
    # price
    #worksheet.set_column('I:K', 9, currency_format)
    worksheet.set_column('I:I', 9, currency_format)
    worksheet.set_column('K:K', 8, currency_format_with_penny)
    worksheet.set_column('L:L', 8, currency_format)
    worksheet.set_column('W:X', 9, currency_format)
    #
    red = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006'})
    green = workbook.add_format({'bg_color': '#C6EFCE',
                                 'font_color': '#006100', 'bold':True})

    worksheet.conditional_format('Q1:Q1000', {'type': 'text',
                                                    'criteria': 'containsText',
                                                    'value': "PP",
                                                    'format': green})
    worksheet.conditional_format('I1:I1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 1,
                                                    'maximum': 10000,
                                                    'format': green})
    worksheet.conditional_format('J1:K1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 0.1,
                                                    'maximum': 10,
                                                    'format': green})
    worksheet.conditional_format('F1:F1000', {'type': 'time_period',
                                             'criteria': 'next week',
                                             'format': red})
    worksheet.conditional_format('N1:N1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 0.1,
                                                    'maximum': 2999,
                                             'format': red})
    worksheet.conditional_format('M1:M1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 0.1,
                                                    'maximum': 999,
                                             'format': red})
    worksheet.conditional_format('X1:X1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': -100000,
                                                    'maximum': -100,
                                             'format': red})
    worksheet.conditional_format('X1:X1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 100,
                                                    'maximum': 100000,
                                             'format': green})
    worksheet.conditional_format('A1:AA1000', {'type': 'no_blanks',
                                               'format': border_format})
    worksheet.conditional_format('P1:P1000', {'type': 'cell',
                                                    'criteria': 'between',
                                                    'minimum': 1,
                                                    'maximum': 100000,
                                             'format': green})
def save_opening_output_file(file_path):
    if path.exists(file_path):
        excelApp = client.Dispatch("Excel.Application")
        book = excelApp.Workbooks.open(path.abspath(file_path))
        excelApp.DisplayAlerts = False
        book.Save()
        book.Close()

def export_to_xlsx(path_file, out_file):
    with open(path_file, encoding='utf8') as f:
        json_data = json.load(f)
        df = json_normalize(json_data['content'])
        if not df.empty:
            df['Регион'] = df['Регион'].astype(int)
            df = df.sort_values(by=['Регион', 'Цена за кв.м']).reset_index(drop=True)

            try:
                #что бы можно было оставлять открым файл оутпут
                save_opening_output_file(out_file)
            except Exception as _ex:
                print(_ex)

            with pd.ExcelWriter(out_file,  engine='xlsxwriter', mode="w") as writer:
                df.to_excel(writer,  encoding='utf-8', freeze_panes=(1,1))
                install_setting_of_columns(writer)
        return
