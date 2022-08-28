import numpy as np
import pandas as pd

from sklearn.compose import ColumnTransformer
from sklearn.datasets import fetch_openml
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler, OneHotEncoder, MinMaxScaler
from sklearn.linear_model import Ridge
from sklearn.tree import DecisionTreeRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.metrics import mean_squared_error, mean_absolute_error
import matplotlib.pyplot as plt

np.random.seed(0)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from inc.export_to_xlsx import save_opening_output_file

def set_predicted_in_xls(file_xlsx_path, df):
    print('set_predicted_in_xls')    
    try:
        save_opening_output_file(file_xlsx_path)
    except Exception as _ex:
        print(_ex)
        
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for i, row in df.iterrows():
        ws[f'W{int(row[0]) + 2}'].number_format = '_-* # ##0₽_-;-* # ##0₽-'
        ws[f'W{int(row[0]) + 2}'] = row["predicted"]
        
        ws[f'X{int(row[0]) + 2}'].number_format = '_-* # ##0₽_-;-* # ##0₽-'
        ws[f'X{int(row[0]) + 2}'] = row["predicted"]-row["Цена за кв.м"]
        if row["predicted"]-row["Цена за кв.м"] > 100:
            ws[f'X{int(row[0]) + 2}'].font = Font(bold=True, color='006100')
            ws[f'X{int(row[0]) + 2}'].fill = PatternFill("solid", fgColor="C6EFCE")
            ws[f'P{int(row[0]) + 2}'] = ((row["predicted"]-row["Цена за кв.м"])*row['Общая площадь']/row["Цена"])
            ws[f'P{int(row[0]) + 2}'].number_format = '0%'
            ws[f'P{int(row[0]) + 2}'].font = Font(bold=True, color='006100')
            ws[f'P{int(row[0]) + 2}'].fill = PatternFill("solid", fgColor="C6EFCE")
        elif row["predicted"]-row["Цена за кв.м"] < 0:
            ws[f'X{int(row[0]) + 2}'].font  = Font(color="9C0006")
            ws[f'X{int(row[0]) + 2}'].fill  = PatternFill("solid", fgColor="FFC7CE")

    wb.save(file_xlsx_path)
    return

#удаляем выбросы
def remove_row_emissions(data, row_name):
    print("remove emissions ", row_name)
    lower_bound = data[row_name].quantile(q=0.025)
    upper_bound = data[row_name].quantile(q=0.975)
    #print(data)
    data[(data[row_name] < lower_bound) | (data[row_name] > upper_bound)] = np.nan
    data.dropna(subset=[row_name], inplace=True)

# разделяем регионы, которые сильно расходятся в ценовой политике
def split_df_into_price_region(df):
    high = [78, 77]
    middle = [50]
    df_high_price = df[df["Регион"].isin(set(high))].copy()

    df_middle_price = df[
        (((df["Жителей в нп"] > 1000000) & (df["Жителей в нп"] < 5000000)) | (df["Регион"].isin(set(middle)))) & ~(
            df["Регион"].isin(set(high)))].copy()
    df_low_price = df[
        (df["Жителей в нп"] <= 1000000) & (df["Жителей в нп"] >= 100) & ~(
            df["Регион"].isin(set(high + middle)))].copy()
    return df_high_price, df_middle_price, df_low_price


# разделяем регионы, которые сильно расходятся в ценовой политике
def loading_data(paths):
    # columns = ["Цена за кв.м", "Регион", "Форма проведения", "Жителей в округе",
    #  "Коммерческих объектов", "Жителей h3", "Имущество", "Расстояние до почты", "Этаж"]
    numeric_features = ["Жителей", "Жителей в нп", "Коммерческих объектов", "Общая площадь",
                        "Отдельный вход", "Культурное наследие", "Ремонт", "Этаж"]
    df = pd.DataFrame()
    if type(paths) == list:
        for path in paths:
            df = pd.concat((df, pd.read_excel(path),), axis=0)
    else:
        df = pd.read_excel(paths)

    df['Тип объекта'] = df['Тип объекта'].astype("string")

    for name in numeric_features:
        df[name].replace('None', 0, inplace=True)
        df[name].replace(np.nan, 0, inplace=True)

    df['Жителей в нп'] = df['Жителей в нп'].astype(int)
    df.dropna(subset=['Жителей','Жителей в нп'], inplace=True)
    # df = df.dropna()

    return df

def get_predicted(out_file, training_data = ['torgi/output_archive.xlsx', 'torgi/output_train.xlsx']):
    full_df = loading_data(training_data) #, 'torgi/output_train.xlsx']
    #print(pd.isnull(full_df).any())
    #print(full_df)

    target = "Цена за кв.м"
    numeric_features = ["Жителей", "Жителей в нп", "Коммерческих объектов", "Общая площадь", "Отдельный вход", "Культурное наследие", "Ремонт"]
    categorical_features = ["Регион", "Форма проведения", "Имущество", "Этаж", "Тип объекта"]
    columns = [target] +numeric_features + categorical_features

    # Сброс ограничений на количество символов в записи
    # pd.set_option('display.max_colwidth', None)
    # pd.set_option('display.max_columns', 10)

    #смотрим в каких регионах сильный разброс цены за кв.м
    full_df.boxplot(by ='Регион', column =[target], grid = False, figsize=(20,10))


    df_high, df_middle, df_low = split_df_into_price_region(full_df)

    df_high.corrwith(df_high[target])

    df_high[target].hist(bins=100, figsize=(10,8))
    df_middle[target].hist(bins=100, figsize=(10,8))
    df_low[target].hist(bins=100, figsize=(10,8))

    #удаляем выбросы
    remove_row_emissions(df_high, "Цена за кв.м")
    remove_row_emissions(df_middle, "Цена за кв.м")
    remove_row_emissions(df_low, "Цена за кв.м")

    #подготавливаем модель для обучения
    numeric_transformer = Pipeline(
        steps=[("imputer", SimpleImputer(strategy='median')), ("scaler", MinMaxScaler())]
    )
    #SimpleImputer(strategy='constant', fill_value=1)

    categorical_transformer = OneHotEncoder(handle_unknown="ignore")

    preprocessor = ColumnTransformer(
        transformers=[
            ("num", numeric_transformer, numeric_features),
            ("cat", categorical_transformer, categorical_features),
        ]
    )

    clf_h = Pipeline(
        steps=[("preprocessor", preprocessor), ("regressor", DecisionTreeRegressor(max_depth=4))]
    )

    df_test_h, df_test_m, df_test_l = split_df_into_price_region(loading_data(out_file))

    # X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)
    if not df_test_h.empty:
        #тренируем модель
        print(df_high.dtypes)
        y_h = df_high["Цена за кв.м"].values
        X_h = df_high[numeric_features + categorical_features]
        clf_h.fit(X_h, y_h)

        y_test_h = df_test_h["Цена за кв.м"].values
        X_test_h = df_test_h[numeric_features + categorical_features]

        #Для модели не достаточно данных для обучения, что бы научится адекватно предсказывать цену в этих регионах
        print("model train score: %.3f" % clf_h.score(X_h, y_h))
        print("model test score: %.3f" % clf_h.score(X_test_h, y_test_h))
        print("MAE: %.3f" % mean_absolute_error(y_test_h, clf_h.predict(X_test_h)))
        print("MAE: %.3f" % mean_absolute_error(y_h, clf_h.predict(X_h)))

        # смотрим результаты для Москвы и Питера
        df_test_h['predicted'] = clf_h.predict(df_test_h[numeric_features + categorical_features])
        print(df_test_h[["Цена за кв.м", 'predicted']].head())
        # очень большой разброс, которым нельзя доверять
        set_predicted_in_xls(out_file, df_test_h)

    if not df_test_m.empty:
        clf_m = Pipeline(
            steps=[("preprocessor", preprocessor), ("regressor", DecisionTreeRegressor(max_depth=4))]
        )

        y_m = df_middle["Цена за кв.м"].values
        X_m = df_middle[numeric_features + categorical_features]
        clf_m.fit(X_m, y_m)

        y_test_m = df_test_m["Цена за кв.м"].values
        X_test_m = df_test_m[numeric_features + categorical_features]

        print("model train score: %.3f" % clf_m.score(X_m, y_m))
        print("model test score: %.3f" % clf_m.score(X_test_m, y_test_m))
        print("MAE: %.3f" % mean_absolute_error(y_test_m, clf_m.predict(X_test_m)))
        #очень сильные результаты можно на них упираться при принятии решения

        # смотрим результаты для МО ЛО Татарстан, Башкирия, Екат [2, 16, 50, 59, 47]
        df_test_m['predicted'] = clf_m.predict(df_test_m[numeric_features + categorical_features])
        print(df_test_m[["Цена за кв.м", 'predicted']].head())
        set_predicted_in_xls(out_file, df_test_m)

    if not df_test_l.empty:
        clf_l = Pipeline(
            steps=[("preprocessor", preprocessor), ("regressor", DecisionTreeRegressor(max_depth=4))]
        )

        y_l = df_low["Цена за кв.м"].values
        X_l = df_low[numeric_features + categorical_features]
        print(X_l.info)
        clf_l.fit(X_l, y_l)

        y_test_l = df_test_l["Цена за кв.м"].values
        X_test_l = df_test_l[numeric_features + categorical_features]

        print("model train score: %.3f" % clf_l.score(X_l, y_l))
        print("model test score: %.3f" % clf_l.score(X_test_l, y_test_l))
        print("MAE: %.3f" % mean_absolute_error(y_test_l, clf_l.predict(X_test_l)))

        #смотрим результаты для всех остальных результатов
        df_test_l['predicted'] = clf_l.predict(df_test_l[numeric_features+categorical_features])
        print(df_test_l[["Цена за кв.м", 'predicted']].head())
        set_predicted_in_xls(out_file, df_test_l)

if __name__ == "__main__":
    get_predicted("../torgi/output.xlsx", "../torgi/output_archive.xlsx")
