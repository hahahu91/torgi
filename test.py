import re
import json
import numpy as np
import pandas as pd
import os
from get_coord import get_location, get_info_object, get_postal_distance
from population_in_district import count_all_objs_in_region, count_entity, get_objs_in_district_from_cache
from h3pandas_test import get_population_from_kontur_population, get_all_objs_from_kontur_population
from win32com import client

from openpyxl import load_workbook
from openpyxl.styles import Font

def save_opening_output_file(file_path):
    if os.path.exists(file_path):
        excelApp = client.Dispatch('Excel.Application')
        book = excelApp.Workbooks.open(os.path.abspath(file_path))
        excelApp.DisplayAlerts = False
        book.Save()
        book.Close()
#

def try_get_coords_again(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        #coord = row[13].value
        address = row[7].value
        cadastr = row[9].value
        population = row[19].value
        id = re.sub(r'^.+"([^"]+)"\)$', r'\1',row[3].value) if row[3].value else None
        lat, lon = get_coord_from_hyper(row[13].value)
        #coord = re.sub(r'^.+"([^"]+)"\)$', r'\1', row[13].value) if row[13].value else None

        if lat and lon and not population:
            postal_distance = get_postal_distance(lat, lon)
            ws[f'T{index + 1}'] = f'{postal_distance}'
            save_population_in_cache(id, postal_distance)


        if not lat or not lon:
            print(f"get info {index} in xls")
            info_object = get_info_object(id, address, cadastr)

            if info_object:
                lat = info_object['lat']
                lon = info_object['lon']
                address = info_object['address']
                coord = f'''=HYPERLINK("https://yandex.ru/maps/?&text={lat}, {lon}", "{lat}, {lon}")''' if lat and lon and lat != "None" and lon != "None" else ""

                ws[f'N{index+1}'] = coord
                ws[f'H{index+1}'] = f'{address}'
                ws[f'T{index+1}'] = f'{info_object["postal_distance"]}'

        elif not os.path.exists(f'cache/tmp_loc/{id}.json'):
            save_address_with_geo(id, lat, lon, address, cadastr)
    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)

def save_population_in_cache(id, postal_distance):
    with open(f'cache/tmp_loc/{id}.json', encoding='utf8') as f:
        info_object = json.load(f)
        info_object[0]['data']['postal_distance'] = postal_distance

        with open(f'cache/tmp_loc/{id}.json', "w", encoding='utf8') as file:
            json.dump(info_object, file, ensure_ascii=False, indent=4)


def save_address_with_geo(id, lat, lon, address, cadnum = None):
    data_json = [{
        "value": f"{address}",
        "unrestricted_value": f"{address}",
        "data": {
            "country": "Россия",
            "country_iso_code": "RU",
            "geo_lat": f"{lat}",
            "geo_lon": f"{lon}",
            "qc_geo": "1",
            "flat_cadnum": f"{cadnum}" if cadnum else None
        }
    }]
    if not os.path.exists('cache/tmp_loc'):
        os.makedirs('cache/tmp_loc')
    with open(f'cache/tmp_loc/{id}.json', "w", encoding='utf8') as file:
        json.dump(data_json, file, ensure_ascii=False, indent=4)

def get_from_kontur_population(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            population_in_h3 = row[17].value
            lat, lon = get_coord_from_hyper(row[13].value)
            if lat and lon and not population_in_h3:
                if not os.path.exists(f'cache/population_in_h3/{lat}_{lon}.json'):
                    objs[index] = {"lat": lat, "lon": lon, "price_m2": row[5].value}
                else:
                    with open(f'cache/population_in_h3/{lat}_{lon}.json', encoding='utf8') as f:
                        info_object = json.load(f)
                        set_in_cell(ws, index=index + 1, population=info_object["population"],
                                    price_m2=info_object["price_m2"])

        except Exception as _ex:
            print(_ex, row[13].value, index)
            raise

    get_all_objs_from_kontur_population(objs)
    for index in objs:
        set_in_cell(ws, index=index+1, population=objs[index]["population"], price_m2=objs[index]["price_m2"])

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)

def get_coord_from_hyper(hyper):
    coord = re.sub(r'^.+"([^"]+)"\)$', r'\1', hyper).split(',') if hyper else None
    if coord:
        return float(coord[0]), float(coord[1])
    return None, None

def set_in_cell(ws, index, population, price_m2):
    ws[f'R{index}'] = population
    ws[f'S{index}'] = float(format(price_m2 / int(population), ".2f")) if int(population) else ""
    ws[f'S{index}'].number_format = '_-* # ##0.00 ₽_-;-* # ##0.00 ₽_-'

def set_population_in_xls_from_cache(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            if row[14].value and row[15].value: #уже есть данные
                continue

            lat, lon = get_coord_from_hyper(row[13].value)
            if not lat or not lon:
                continue

            price_m2 = row[5].value

            if lat and lon:
                if os.path.exists(f'cache/objs_in_district/{lat}_{lon}.json'):
                    entity = get_objs_in_district_from_cache(lat, lon)
                    if entity:
                        ws[f'O{index+1}'] = f'{entity["residents"]}'
                        ws[f'P{index+1}'] = f'{entity["entity"]}'
                        ws[f'Q{index+1}'].font = Font(underline='single', color='0000FF')
                        ws[f'Q{index+1}'] = f"""=HYPERLINK("{os.path.abspath("cache/objs_in_district")}/{lat}_{lon}.json", "{lat}_{lon}.json")"""

                        ws[f'K{index+1}'] = float(format(price_m2 / int(entity["residents"]), ".2f")) if int(entity["residents"]) else ""
                        ws[f'K{index+1}'].number_format = '_-* # ##0.00 ₽_-;-* # ##0.00 ₽_-'

        except Exception as _ex:
            print(_ex)

    save_opening_output_file(file_xlsx_path)
    wb.save(file_xlsx_path)

def get_population_from_osm(file_xlsx_path):
    wb = load_workbook(file_xlsx_path)
    ws = wb['Sheet1']
    objs = {}
    objs_by_regions = {}
    for index, row in enumerate(ws.rows):
        if (index == 0):
            continue
        try:
            region = row[1].value
           # objs_by_regions.update(region)
            if row[13].value == 'None, None' or row[13].value == None or row[13].value == '':
                continue

            coord_lat, coord_lon =re.sub(r'^.+"([^"]+)"\)$', r'\1', row[13].value).split(',') if row[13].value else None
            lat, lon = float(coord_lat), float(coord_lon) or None
            price_m2 = row[5].value
            if lat and lon:
                if not os.path.exists(f'cache/objs_in_district/{lat}_{lon}.json'):
                    objs[index] = {"index": index, "lat": lat, "lon": lon, "price_m2": price_m2, 'nodes':set(), 'entity':{}}
                    list_obj = objs_by_regions.get(region) or []
                    list_obj.insert(index,objs[index])
                    objs_by_regions[region] = list_obj #{objs_by_regions[region], {index:objs[index]}}

        except Exception as _ex:
            print(_ex, row[13].value, index)
            raise

    for region in objs_by_regions:
        print(objs_by_regions[region])
        count_all_objs_in_region(objs_by_regions[region], region)

    set_population_in_xls_from_cache(file_xlsx_path)

    return

#try_get_coords_again('torgi/output_archive.xlsx')
get_from_kontur_population('torgi/output_archive.xlsx')
get_population_from_osm('torgi/output_archive.xlsx')

